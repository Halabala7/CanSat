
import matplotlib.pyplot as plt
import serial
import folium
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from decimal import Decimal


def draw_graph(number_of_measurements, pressure, temperature, altitude_from_bmp, altitude_from_gps, latency):

    def plt_graphs(first_data, second_data, title, first_name, second_name, save_name):

        plt.figure(figsize=(15, 10))
        plt.plot(first_data, second_data)

        plt.title(title)
        plt.xlabel(first_name)
        plt.ylabel(second_name)
        plt.axis([min(first_data), max(first_data),
                  min(second_data), max(second_data)])
        plt.savefig(save_name)
        plt.close()
    plt_graphs(number_of_measurements, temperature, 'wykres temperatury względem liczby pomiarow',
               'pomiar', 'temperatura', '../dane/temperatura_z_pomiarami.png')
    plt_graphs(number_of_measurements, pressure, 'wykres cisnienia względem liczby pomiarow',
               'pomiar', 'cisnienie', '../dane/cisnienie_z_pomiarami.png')

    # Tworzenie nowego arkusza Excela
    wb = Workbook()
    ws_temp_gps = wb.active  # Zakładka dla temperatury z GPS
    ws_temp_gps.title = "Temperature_GPS"
    # Utworzenie nowej zakładki dla ciśnienia z GPS
    ws_pres_gps = wb.create_sheet("Pressure_GPS")
    # Utworzenie nowej zakładki dla temperatury z czujnika bmp280
    ws_temp_bmp = wb.create_sheet("Temperature_BMP")
    # Utworzenie nowej zakładki dla ciśnienia z czujnika bmp280
    ws_pres_bmp = wb.create_sheet("Pressure_BMP")

    ws_lat_bmp = wb.create_sheet("Latency_BMP")

    # Dane do wykresów
    temperature = [Decimal(str(temp)).quantize(Decimal('0.00'))
                   for temp in temperature]
    pressure = [Decimal(str(temp)).quantize(Decimal('0.00'))
                for temp in pressure]
    altitude_from_bmp = [Decimal(str(temp)).quantize(Decimal('0.00'))
                         for temp in altitude_from_bmp]
    altitude_from_gps = [Decimal(str(temp)).quantize(Decimal('0.00'))
                         for temp in altitude_from_gps]
    latency = [Decimal(str(temp)).quantize(Decimal('0.00'))
               for temp in latency]

    # Dodawanie początkowych danych do arkusza dla temperatury i ciśnienia z GPS
    ws_temp_gps.append(['Altitude from gps', 'Temperature'])
    for alt, temp in zip(altitude_from_gps, temperature):
        ws_temp_gps.append([alt, temp])

    ws_pres_gps.append(['Altitude from gps', 'Pressure'])
    for alt, pres in zip(altitude_from_gps, pressure):
        ws_pres_gps.append([alt, pres])

    # Dodawanie początkowych danych do arkusza dla temperatury i ciśnienia z czujnika bmp280
    ws_temp_bmp.append(['Altitude from bmp', 'Temperature'])
    for alt, temp in zip(altitude_from_bmp, temperature):
        ws_temp_bmp.append([alt, temp])

    ws_pres_bmp.append(['Altitude from bmp', 'Pressure'])
    for alt, pres in zip(altitude_from_bmp, pressure):
        ws_pres_bmp.append([alt, pres])

    ws_lat_bmp.append(['Altitude from bmp', 'Latency'])
    for alt, lat in zip(altitude_from_bmp, latency):
        ws_lat_bmp.append([alt, lat])

    # Tworzenie wykresów dla temperatury i ciśnienia z GPS
    def create_chart(ws, title, data, categories, chart_cell):
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = 'Altitude'
        chart.y_axis.title = title.split()[0]
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 40  # Szerokość wykresu
        chart.height = 32  # Wysokość wykresu
        ws.add_chart(chart, chart_cell)

    # Tworzenie wykresu dla temperatury z GPS
    data_temp_gps = Reference(ws_temp_gps, min_col=2, min_row=1,
                              max_col=2, max_row=len(temperature) + 1)
    cats_temp_gps = Reference(ws_temp_gps, min_col=1, min_row=2,
                              max_row=len(temperature) + 1)
    create_chart(ws_temp_gps, "Temperature vs. Altitude (GPS)",
                 data_temp_gps, cats_temp_gps, "D1")

    # Tworzenie wykresu dla ciśnienia z GPS
    data_pres_gps = Reference(ws_pres_gps, min_col=2, min_row=1,
                              max_col=2, max_row=len(pressure) + 1)
    cats_pres_gps = Reference(ws_pres_gps, min_col=1, min_row=2,
                              max_row=len(pressure) + 1)
    create_chart(ws_pres_gps, "Pressure vs. Altitude (GPS)",
                 data_pres_gps, cats_pres_gps, "D1")

    # Tworzenie wykresu dla temperatury z czujnika bmp280
    data_temp_bmp = Reference(ws_temp_bmp, min_col=2, min_row=1,
                              max_col=2, max_row=len(temperature) + 1)
    cats_temp_bmp = Reference(ws_temp_bmp, min_col=1, min_row=2,
                              max_row=len(temperature) + 1)
    create_chart(ws_temp_bmp, "Temperature vs. Altitude (BMP280)",
                 data_temp_bmp, cats_temp_bmp, "D1")

    # Tworzenie wykresu dla ciśnienia z czujnika bmp280
    data_pres_bmp = Reference(ws_pres_bmp, min_col=2, min_row=1,
                              max_col=2, max_row=len(pressure) + 1)
    cats_pres_bmp = Reference(ws_pres_bmp, min_col=1, min_row=2,
                              max_row=len(pressure) + 1)
    create_chart(ws_pres_bmp, "Pressure vs. Altitude (BMP280)",
                 data_pres_bmp, cats_pres_bmp, "D1")

    data_latency = Reference(ws_lat_bmp, min_col=2, min_row=1,
                             max_col=2, max_row=len(latency) + 1)
    cats_pres_bmp = Reference(ws_lat_bmp, min_col=1, min_row=2,
                              max_row=len(pressure) + 1)
    create_chart(ws_lat_bmp, "Latency vs. Altitude (BMP280)",
                 data_latency, cats_pres_bmp, "D1")

    # Zapisanie pliku Excela
    wb.save("../dane/dane_i_wykres.xlsx")


def save_data_to_files(number_of_measurements, pressure, temperature, altitude_from_bmp, latitude, longitude, altitude_from_gps, latency):
    file_temperature = open("../dane/temperatura.txt", "w")
    file_pressure = open("../dane/cisnienie.txt", "w")
    file_lat_long = open("../dane/wspolrzedne.txt", "w")
    file_latency = open("../dane/opoznienie.txt", "w")

    avg = round((sum(list(map(int, latency)))/len(latency)), 2)
    file_latency.seek(0)
    file_latency.writelines("max:" + max(latency) +
                            "min: " + min(latency) + "avg: " + str(avg) + "\n")

    if latitude != 0 or longitude != 0 or altitude_from_gps != 0:
        latitude = list(map(float, latitude))
        longitude = list(map(float, longitude))

    for i in range(len(temperature)):

        if latitude != 0 or longitude != 0 or altitude_from_gps != 0:

            file_temperature.write("temperatura: " + str(temperature[i]) + "°C, pomiar: " + str(
                number_of_measurements[i]) + ", na wysokosci (cisnienie): " + str(altitude_from_bmp[i]) + " (gps): " + str(altitude_from_gps[i]) + " m.n.p.m" + "\n")
            file_pressure.write("cisnienie: " + str(pressure[i]) + " hPa, pomiar: " + str(
                number_of_measurements[i]) + ", na wysokosci (cisnienie): " + str(altitude_from_bmp[i]) + " (gps): " + str(altitude_from_gps[i]) + " m.n.p.m" + "\n")
            file_lat_long.write("szerokosc geograficzna: " +
                                str(latitude[i]) + ", dlugosc geograficzna: " + str(longitude[i]) + "\n")

        else:
            file_temperature.write("temperatura: " + str(temperature[i]) + "°C, pomiar: " + str(
                number_of_measurements[i]) + ", na wysokosci: " + str(altitude_from_gps[i]) + " brak danych " + " m.n.p.m" + "\n")
            file_pressure.write("cisnienie: " + str(pressure[i]) + "hPa, pomiar: " + str(
                number_of_measurements[i]) + ", na wysokosci: " + str(altitude_from_gps[i]) + " brak danych " + " m.n.p.m" + "\n")
            file_lat_long.write("szerokosc geograficzna: " +
                                latitude + ", dlugosc geograficzna: " + longitude + " brak danych " + "\n")

    file_temperature.close()
    file_pressure.close()
    file_lat_long.close()
    file_latency.close()


def save_raw_data(latency,temperature,pressure,altitude_from_bmp, latitude, longitude, altitude_from_gps):
    file_raw_data = open("../dane/surowe_dane.txt", "w")
    for i in range(len(latency)):
        file_raw_data.write(str(latency[i]) + "," + str(temperature[i]) + "," + str(pressure[i]) + "," + str(altitude_from_bmp[i]) + "," + str(latitude[i]) + "," + str(longitude[i]) + "," + str(altitude_from_gps[i]) + "," + "\n")
    file_raw_data.close()


def show_gps(latitude, longitude):

    map = folium.Map(location=[latitude[0], longitude[0]], zoom_start=20)

    for i in range(len(latitude) - 1):
        popup_cont = f"Position: {latitude[i]} {longitude[i]}"
        folium.CircleMarker([latitude[i], longitude[i]], popup=popup_cont,
                            radius=2, color='blue', fill=True, fill_color='lightblue').add_to(map)

    lat = latitude[-1]
    lon = longitude[-1]
    popup_content = f"Current Position: {lat} {lon}"

    folium.Marker([lat, lon], popup=popup_content, icon=folium.Icon(
        color='orange', icon='map-marker')).add_to(map)

    folium.PolyLine(list(zip(latitude, longitude)), color="orange").add_to(map)

    map.save('../dane/map.html')


ser = serial.Serial('COM5', 9600)

latency = []
temperature = []
pressure = []
altitude_from_bmp = []
latitude = []
longitude = []
altitude_from_gps = []
number_of_measurements = []

counter = 0
measurement = 1

while True:

    if ser.in_waiting > 0:
        data = ser.readline().decode('utf-8').strip()

        print(data)

        text = ""
        dataList = []
        for letter in data:

            if letter == ",":
                dataList.append(text)
                text = ""
            else:
                text = text + letter
        if len(dataList) == 3:
            latency.append(dataList[0])
            temperature.append(dataList[1])
            pressure.append(dataList[2])
            altitude_from_bmp.append(dataList[3])
            save_raw_data(latency,temperature,pressure,altitude_from_bmp, "0", "0", "0")
        else:
            latency.append(dataList[0])
            temperature.append(dataList[1])
            pressure.append(dataList[2])
            altitude_from_bmp.append(dataList[3])
            latitude.append(float(dataList[4]))
            longitude.append(float(dataList[5]))
            altitude_from_gps.append(dataList[6])
            save_raw_data(latency,temperature,pressure,altitude_from_bmp, latitude, longitude, altitude_from_gps)
        number_of_measurements.append(measurement)
        measurement += 1
        counter += 1

        if len(dataList) != 3:
            show_gps(latitude, longitude)

        if counter == 10:
            counter = 0

            if len(dataList) == 3:
                save_data_to_files(number_of_measurements,
                                   pressure, temperature, " ", " ", " ")
                print("brak dancyh gps, wyslano niepelne dane")
            else:
                draw_graph(number_of_measurements, pressure, temperature,
                           altitude_from_bmp, altitude_from_gps, latency)
                save_data_to_files(number_of_measurements, pressure, temperature, altitude_from_bmp,
                                   latitude, longitude, altitude_from_gps, latency)
                print("wypisano dane, sprawdz pliki")
